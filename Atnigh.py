import os,re,time,csv,requests
import xlwt,jsonpath
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl import workbook
from openpyxl.utils import get_column_letter
import numpy as np
import pandas as pd
from time import sleep
from lxml import etree
from pprint import pprint
from atnight_baidu_nlp import BaiDuAI as AI
from retrying import retry
from PIL import Image
import pytesseract
import cv2,requests
from urllib import request


def forFileName(path):
    #提取文件夹内文件名
    for a,b,c in os.walk(path):
        return a,b,c


def get_all_path(open_file_path):
    #获取所有文件路径
    rootdir = open_file_path
    path_list = []
    list = os.listdir(rootdir)  # 列出文件夹下所有的目录与文件
    try:
        for i in range(0,len(list)):
            com_path = os.path.join(rootdir,list[i])
            #print(com_path)
            if os.path.isfile(com_path):
                path_list.append(com_path)
            if os.path.isdir(com_path):
                path_list.extend(get_all_path(com_path))
    except Exception as result:print(result)
    return path_list


def parse_id(url):
    #正则提取ID，消除所有标点符号
    partten = re.compile(r"id=(\d+)")
    partten = partten.findall(url)
    return "".join(partten)

def forCombination():
    # 列出列表所有组合
    path = r"G:\testfiles\atnight\test.xlsx"
    wb = load_workbook(filename=path)
    sheet1 = wb.active

    scenario = []
    painPoints = []
    demand = []
    for i in sheet1["A"]:
        scenario.append(i.value)
    for i in sheet1["B"]:
        painPoints.append(i.value)
    for i in sheet1["C"]:
        demand.append(i.value)

    scenario.remove("场景")
    painPoints.remove("痛点")
    demand.remove("需求")
    pprint(scenario)
    pprint(painPoints)
    pprint(demand)


def folderToExtractId(path,switchTag=0):
    #提取概况数据中的id和店铺名
    a,b,c = forFileName(path)
    for x in b:
        x = x.split("_")
        try:
            link = linksToCompletion(x[1].replace("ID",""))
            if switchTag==1:
                print(x[2],x[1].replace("ID",""),link)
            else:
                print(x[2],x[1].replace("ID",""))
        except:
            pass

def toDealWithSimilarNumerical(path,floatingRange,priceFloor,priceCeil):
    # 消除价格浮动 按众数
    wb = load_workbook(filename=path)
    sheet1 = wb.active
    for i in sheet1["D"]:
        try:
            if float(i.value)>priceFloor and float(i.value)<priceCeil:
                    priceList = []
                    for i2 in sheet1["D"]:
                        try:
                            x = float(i.value)-float(i2.value)
                            if abs(x)<=floatingRange:
                                priceList.append(float(i2.value))
                        except:
                            pass
                    #print(priceList)
                    #print(len(priceList))
                    #i.value = max(priceList)
                    if len(priceList)>0:
                        i.value = publicnum(priceList)
                    else:
                        pass
        except:
                pass
    wb.save(filename=path)

def publicnum(num, d = 0):
    # 传入列表求众数
    dictnum = {}
    for i in range(len(num)):
        if num[i] in dictnum.keys():
            dictnum[num[i]] += 1
        else:
            dictnum.setdefault(num[i], 1)
    maxnum = 0
    maxkey = 0
    for k, v in dictnum.items():
        if v > maxnum:
            maxnum = v
            maxkey = k
    return maxkey

def sentimentAnalysis(sheet):
    for i in sheet["C2:C{}".format(sheet.max_row)]:
        for j in i:
            j.value = AI().filter_emoji(desstr=j.value)
            toParseJson = AI().run1(content=j.value)
            sleep(0.5)
            # 解析json数据
            toParseJson_sentiment = jsonpath.jsonpath(toParseJson,'$..sentiment')
            toParseJson_abstract = jsonpath.jsonpath(toParseJson,'$..abstract')
            toParseJson_prop = jsonpath.jsonpath(toParseJson,'$..prop')
            toParseJson_adj = jsonpath.jsonpath(toParseJson,'$..adj')
            print(toParseJson_sentiment,toParseJson_abstract,toParseJson_prop,toParseJson_adj)
            sheet["D{}".format(j.row)].value = str(toParseJson_sentiment) if isinstance(toParseJson_sentiment,list) else "none"
            if sheet["D{}".format(j.row)].value.find("0")==-1:
                sheet.delete_rows(j.row,1)

def aboutMerger(path,recognitionField,selectSchema=2):
    # 处理评论和问大家和SKU总览的xlsx文件
    # 取文件路径
    a,b,c = forFileName(path)
    for x in tqdm(b):
        # 需要文件夹名称规范，否则需要以另外方式取得名字
        shopNameId = x.split("_")[2]
        shopNameId2 = x.split("_")[1].replace("ID","")
        a1,b1,c1 = forFileName(os.path.join(a,x))

        for x2 in c1:
            if x2.find(recognitionField)!=-1:
                askEveryoneComment = os.path.join(a,x,x2)
                print(askEveryoneComment)  # 打印所有符合条件文件的路径
                # 读写Excel格式文件处理数据
                try:
                    wb = load_workbook(askEveryoneComment)
                    sheet = wb.active
                    sheet.insert_cols(1,2)
                    sheet["A1"].value = "店铺名"
                    sheet["B1"].value = "产品ID"
                    i = 2
                    # 第一、第二列加上店铺和id的定位标识
                    for x3 in sheet["A2:A{}".format(sheet.max_row)]:
                            for x4 in x3:
                                x4.value = shopNameId
                    for x3 in sheet["B2:B{}".format(sheet.max_row)]:
                        for x4 in x3:
                            x4.value = shopNameId2

                    if selectSchema == 2:  # 参数2表示要做这个评论分析，改为其他参数就跳过

                        # 单独给评论文件做情感分析，首先通过文件名筛选评论文件，.find()!=-1表示找到了
                        if askEveryoneComment.find("评论")!=-1 and sheet.max_row>50:
                            sheet["C1"].value = "首评+追评"
                            sheet["D1"].value = "Sentiment"
                            for x3 in sheet["C2:C{}".format(sheet.max_row)]:
                                for x4 in x3:
                                    x4.value = str(sheet["F{}".format(i)].value) if not None else ""+str(sheet["J{}".format(i)].value) if not None else ""
                                    i+=1
                            sheet.delete_cols(4,sheet.max_column-3)
                            # 删除长度短于6的字符串，还有就是默认的评价，然后去重（去重还没有实现）
                            for x3 in sheet["C2:C{}".format(sheet.max_row)]:
                                for j in x3:
                                    if len(j.value)<6 or j.value.find("此用户没有填写评价")!=-1 or j.value.find("系统默认好评")!=-1:
                                        sheet.delete_rows(j.row,1)
                            # 检查情况对不对
                            #for x3 in sheet["C2:C{}".format(sheet.max_row)]:
                            #    for j in x3:
                            #        print(j.value)
                            countItems = 2
                            countItems2 = 0
                            countItems3 = 2
                            # 这里的10可以设定为超参数控制迭代倍数，从而控制次数
                            for totalNumberCycles in range(1,(sheet.max_row//10)+1):
                                for i in range(3,13):
                                    sheet["C{}".format(countItems)].value = str(sheet["C{}".format(countItems)].value)+"_"+str(sheet["C{}".format(i+countItems2)].value)
                                    sheet["C{}".format(i+countItems2)].value = ""
                                countItems+=1
                                countItems2+=10
                            # 移除空行
                            sheet.delete_rows((sheet.max_row//10)+2,sheet.max_row)
                            # 第一轮NLP-情感分析,删除非0值行
                            sentimentAnalysis(sheet)
                            # 拆开重组后再进行一次NLP，再删除非0值行，得到结果，突然发现，这个的循环是不是就是卷积的过程
                            for i in sheet["C2:C{}".format(sheet.max_row)]:
                                for j in i:
                                    breakUpData = j.value.split("_")
                                    for i2 in breakUpData:
                                        if len(i2)>5:
                                            sheet["D{}".format(countItems3)].value = i2
                                            countItems3+=1
                            # 把拆解结果移回C列
                            sheet.move_range("D2:D{}".format(sheet.max_row),cols=-1)
                            # 这里补上一次NLP-情感分析
                            sentimentAnalysis(sheet)
                            # 补上店铺和id
                            for x3 in sheet["A2:A{}".format(sheet.max_row)]:
                                for x4 in x3:
                                    x4.value = shopNameId
                            for x3 in sheet["B2:B{}".format(sheet.max_row)]:
                                for x4 in x3:
                                    x4.value = shopNameId2
                    wb.save(filename=askEveryoneComment)
                except Exception as result:
                    print(result)

def imageOCR(image):
    # 图片OCR
    content = pytesseract.image_to_string(image, lang='chi_sim')   # 解析图片
    print(content)

def take3CAuthentication():
    # 3C认证,站内可查，不必站外
    data = {
        "keyword":"暖风机",
        "chaxuntype":"产品名称 Product Name",
        "_h_select_chaxuntype":"product",
        "chaxuntype": "product",
        "pageSize":"100",
        "sortColumns":"null",
        "pageNumber":"1"
            }

    headers = {
        "User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36",
        "Host":"webdata.cqccms.com.cn",
        "Origin":"http://webdata.cqccms.com.cn"
        }
    start_url = "http://webdata.cqccms.com.cn/webdata/query/CCCCerti.do"
    session = requests.Session()
    r = session.post(url=start_url,headers=headers,data=data)
    if r.status_code==200:
        verificationCode = "http://webdata.cqccms.com.cn/webdata/ImagepassController/imagePass.do"
        request.urlretrieve(verificationCode,"verificationCode.png")
        img = Image.open(r"verificationCode.png")
        img.show()
        data["imagePassword"] = input("请输入验证码：")
        r2 = session.post(start_url,headers=headers,data=data)
        print(r2.status_code)
        print(r2.text)



if __name__=='__main__':

    # 图像OCR模块
    #image = cv2.imread(r"D:\Testfiles\zhutu.jpg")
    #imageOCR(get_grayscale(image))  # 效果不好

    take3CAuthentication()  # 3C

    #inputTest()

    # 遍历文件夹转换csv为xlsx
    #traversalFolder(r"D:\atnight\蓝禾科技\车用吸尘器\评论分析")

    # 评论文件情感分析
    #aboutMerger(r"E:\","评论分析",2)

    print("-" * 50,"{0}".format(time.strftime('%Y-%m-%d %H:%M')),"-" * 50)