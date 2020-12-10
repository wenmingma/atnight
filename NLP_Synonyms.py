import os,re,time,shutil,csv,requests,jieba,sys,jieba.analyse
import itertools,xlwt,json,jsonpath,synonyms
from tqdm import tqdm,trange
from openpyxl import load_workbook
from openpyxl import workbook
from openpyxl.utils import FORMULAE
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string
from openpyxl.drawing.image import Image
import pandas as pd
import numpy as np
from time import sleep
from jieba import analyse
from pprint import pprint


def participles():
    jieba.load_userdict("dict.txt")
    words=jieba.cut("手机直播补光灯")
    print("/".join(words))
    print(type(words))

def keywordExtraction(path,columnName):
    # jieba分词提取关键词
    jieba.load_userdict("dict.txt")  # 加载字典

    wb = load_workbook(path)
    sheet = wb.active

    # 这个算法好像没tf-idf好,好像
    #textRank = analyse.textrank  

    countItems = 2
    contentInColumn = str(columnName)  # 内容所在列
    contentTheOutputColumns = get_column_letter(column_index_from_string(contentInColumn)+1)
    sheet["{}1".format(contentTheOutputColumns)] = "MachineToExtract"  # 内容输出列
    for i in sheet["{0}2:{1}{2}".format(contentInColumn,contentInColumn,sheet.max_row)]:
        for j in i:

            # 这段提取标签
            #keywords = textRank(j.value)
            try:
                # 到底用什么词性才是最贴切呢？
                keywords = jieba.analyse.extract_tags(j.value,3,allowPOS=('n','v','a'))
                keywordsConvert = " ".join(keywords)
                print(j.value,keywords)
                sheet["{0}{1}".format(contentTheOutputColumns,countItems)].value = keywordsConvert
            except:pass
            countItems+=1
            #for keyword in keywords:
            #    print(j.value,keyword)
            

            # 这段用来比较句子的相似程度，但是词库缺乏词条
            #for i2 in sheet["{}{}:{}{}".format(contentInColumn,j.row+1,contentInColumn,sheet.max_row)]:
            #    for j2 in i2:
            #        r = synonyms.compare(j.value,j2.value)
            #        if r>0.7:
            #            sheet.delete_rows(j2.row,1)
            #print(j.value)

    wb.save(filename=path)

def similarityComparison():
    # 近义词处理，缺少近义词库
    wb = load_workbook(path)
    sheet = wb.active
    for i in sheet["E2:E{}".format(sheet.max_row)]:
        for j in i:
            for i2 in sheet["E3:E{}".format(sheet.max_row)]:
                pass
            r = synonyms.compare(j.value,)

if __name__ =="__main__":

    # 近义词处理
    #aSynonymFor()

    # jieba分词处理（加载分词字典）
    #participles()

    # textrank算法提取关键词
    keywordExtraction(r"G:\testfiles\3支架-问大家(2020-08-18 090733).xlsx","C")

    print("-" * 50,"{0}".format(time.strftime('%Y-%m-%d %H:s%M')),"-" * 50)